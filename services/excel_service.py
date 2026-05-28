from __future__ import annotations

import logging
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from models.asset_record import AssetRecord
from utils.constants import (
    APP_NAME,
    EXCEL_SHEET_SOURCE,
    LABEL_IMAGE_HEIGHT_PX,
    LABEL_IMAGE_WIDTH_PX,
)
from utils.normalization import normalize_excel_scalar

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError:  # pragma: no cover
    Image = None
    ImageDraw = None
    ImageFont = None


@dataclass(slots=True)
class LabelImageExportResult:
    position: int
    group_name: str
    output_path: Path
    original_size_px: tuple[int, int]
    target_size_px: tuple[int, int]


class ExcelService:
    """Headless Excel reader + stable label renderer.

    This service intentionally avoids Microsoft Excel COM entirely.
    It reads workbook data with openpyxl and renders each label image
    to match the visual layout of the real `Etiqueta provisional` design.
    """

    def __init__(self) -> None:
        self.logger = logging.getLogger(APP_NAME)
        self.workbook = None
        self.source_sheet = None
        self._source_headers: list[str] = []
        self._records_cache: list[AssetRecord] = []
        self._current_block: list[AssetRecord] = []
        self._logo_image = None

    def open(self, workbook_path: str, visible: bool = False) -> None:  # noqa: ARG002
        workbook_file = Path(workbook_path).resolve()
        suffix = workbook_file.suffix.lower()
        if suffix not in {".xlsx", ".xlsm"}:
            raise RuntimeError("Solo se admiten archivos .xlsx o .xlsm para lectura interna sin Microsoft Excel.")

        self.close(save_changes=False)
        self.workbook = load_workbook(filename=str(workbook_file), data_only=True, read_only=False)
        try:
            self.source_sheet = self.workbook[EXCEL_SHEET_SOURCE]
        except KeyError as exc:
            self.close(save_changes=False)
            raise RuntimeError(f"No existe la hoja fuente requerida: {EXCEL_SHEET_SOURCE}") from exc

        self._source_headers = [
            normalize_excel_scalar(self.source_sheet.cell(1, column).value)
            for column in range(1, self.source_sheet.max_column + 1)
        ]
        self._records_cache = []
        self._current_block = []
        self.logger.info("Excel cargado en modo interno/headless con openpyxl: %s", workbook_file)

    def close(self, save_changes: bool = False) -> None:  # noqa: ARG002
        workbook = self.workbook
        self.workbook = None
        self.source_sheet = None
        self._source_headers = []
        self._records_cache = []
        self._current_block = []
        self._logo_image = None
        if workbook is not None:
            workbook.close()

    def get_sheet_names(self) -> list[str]:
        if self.workbook is None:
            return []
        return list(self.workbook.sheetnames)

    def get_headers(self) -> list[str]:
        return [header for header in self._source_headers if header]

    def get_filters(self) -> list[str]:
        worksheet = self._require_source_sheet()
        values: set[str] = set()
        for row in range(2, worksheet.max_row + 1):
            normalized = self._normalize_excel_text(worksheet.cell(row, 3).value)
            if normalized:
                values.add(normalized)
        return sorted(values)

    def get_filtered_records(self, selected_filter: str) -> list[AssetRecord]:
        worksheet = self._require_source_sheet()
        records: list[AssetRecord] = []
        for row in range(2, worksheet.max_row + 1):
            section_value = self._normalize_excel_text(worksheet.cell(row, 3).value)
            if section_value != selected_filter:
                continue

            records.append(
                AssetRecord(
                    row_index=row,
                    asset_id=self._normalize_excel_identifier(worksheet.cell(row, 1).value),
                    asset_name=self._normalize_excel_text(worksheet.cell(row, 2).value),
                    section=section_value,
                )
            )

        self._records_cache = records
        return records

    def normalize_identifier_column(self) -> list[tuple[int, str, str]]:
        worksheet = self._require_source_sheet()
        audit: list[tuple[int, str, str]] = []
        for row in range(2, worksheet.max_row + 1):
            raw_value = worksheet.cell(row, 1).value
            raw_text = "" if raw_value is None else str(raw_value)
            normalized = self._normalize_excel_identifier(raw_value)
            if raw_text.strip() != normalized:
                audit.append((row, raw_text, normalized))
        return audit

    def write_block_to_label_sheet(self, records: list[AssetRecord]) -> None:
        self._current_block = list(records)

    def clear_label_input_area(self) -> None:
        self._current_block = []

    def get_generated_assets(self, count: int) -> list[str]:
        return [record.asset_id for record in self._current_block[:count]]

    def export_label_shape_image(
        self,
        position: int,
        output_png: str | Path,
        target_px: tuple[int, int] = (LABEL_IMAGE_WIDTH_PX, LABEL_IMAGE_HEIGHT_PX),
    ) -> LabelImageExportResult:
        if Image is None or ImageDraw is None or ImageFont is None:
            raise RuntimeError("Pillow no está instalado. Instalá las dependencias desde requirements.txt.")

        if position < 0 or position >= len(self._current_block):
            raise RuntimeError(f"No hay registro cargado para la posición {position + 1}.")

        record = self._current_block[position]
        image = self._render_record_label(record, target_px)
        output_path = Path(output_png)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        image.save(output_path, format="PNG")

        return LabelImageExportResult(
            position=position + 1,
            group_name=f"record-{position + 1:03d}",
            output_path=output_path,
            original_size_px=image.size,
            target_size_px=target_px,
        )

    def _render_record_label(self, record: AssetRecord, target_px: tuple[int, int]):
        width_px, height_px = target_px
        image = Image.new("RGBA", target_px, (255, 255, 255, 255))
        draw = ImageDraw.Draw(image)

        border_color = (220, 220, 220, 255)
        text_color = (0, 0, 0, 255)

        draw.rectangle((0, 0, width_px - 1, height_px - 1), outline=border_color, width=1)

        margin = 6
        logo_box = (4, 8, 106, 108)
        self._paste_real_logo(image, logo_box)

        section_font = self._load_font(24, bold=False)
        asset_id_font = self._load_font(31, bold=True)
        name_font = self._load_font(20, bold=False)

        section_text = record.section or "SIN SECCIÓN"
        self._draw_centered_single_line(draw, section_text, (112, 8, width_px - 10, 56), section_font, text_color)

        asset_id_text = record.asset_id or "SIN ACTIVO"
        self._draw_centered_single_line(draw, asset_id_text, (112, 52, width_px - 10, 103), asset_id_font, text_color)

        asset_name_text = record.asset_name or "SIN DESCRIPCIÓN"
        self._draw_wrapped_centered_text(draw, asset_name_text, (margin, 112, width_px - margin, height_px - 8), name_font, text_color)

        return image.convert("RGB")

    def _paste_real_logo(self, image, box: tuple[int, int, int, int]) -> None:
        logo = self._load_logo_image()
        if logo is None:
            self._draw_logo_fallback(image, box)
            return

        target_width = max(1, box[2] - box[0])
        target_height = max(1, box[3] - box[1])
        resized = logo.resize((target_width, target_height), self._pil_resize_filter())
        image.alpha_composite(resized, (box[0], box[1]))

    def _load_logo_image(self):
        if self._logo_image is not None:
            return self._logo_image

        if Image is None:
            return None

        logo_path = Path(__file__).resolve().parents[1] / "assets" / "weg_logo.png"
        try:
            self._logo_image = Image.open(logo_path).convert("RGBA")
        except Exception:
            self._logo_image = None
        return self._logo_image

    def _draw_logo_fallback(self, image, box: tuple[int, int, int, int]) -> None:
        draw = ImageDraw.Draw(image)
        draw.rectangle(box, outline=(0, 0, 0, 255), width=3)
        inner_margin = 8
        draw.rectangle(
            (box[0] + inner_margin, box[1] + inner_margin, box[2] - inner_margin, box[3] - inner_margin),
            outline=(0, 0, 0, 255),
            width=3,
        )
        logo_font = self._load_font(max(20, int((box[3] - box[1]) * 0.42)), bold=True)
        self._draw_centered_single_line(draw, "WEG", box, logo_font, (0, 0, 0, 255))

    def _draw_centered_single_line(self, draw, text: str, box: tuple[int, int, int, int], font, fill) -> None:
        left, top, right, bottom = box
        bbox = draw.textbbox((0, 0), text, font=font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        x = left + max(0, ((right - left) - text_width) // 2)
        y = top + max(0, ((bottom - top) - text_height) // 2) - bbox[1]
        draw.text((x, y), text, font=font, fill=fill)

    def _draw_wrapped_centered_text(self, draw, text: str, box: tuple[int, int, int, int], font, fill) -> None:
        left, top, right, bottom = box
        available_width = max(40, right - left)
        available_height = max(20, bottom - top)

        wrapped_lines = self._wrap_text(draw, text, font, available_width)
        while self._text_block_height(draw, wrapped_lines, font) > available_height and getattr(font, "size", 8) > 12:
            font = self._load_font(getattr(font, "size", 8) - 1, bold=False)
            wrapped_lines = self._wrap_text(draw, text, font, available_width)

        total_height = self._text_block_height(draw, wrapped_lines, font)
        current_y = top + max(0, (available_height - total_height) // 2)
        for line in wrapped_lines:
            bbox = draw.textbbox((0, 0), line, font=font)
            line_width = bbox[2] - bbox[0]
            line_height = bbox[3] - bbox[1]
            x = left + max(0, (available_width - line_width) // 2)
            y = current_y - bbox[1]
            draw.text((x, y), line, font=font, fill=fill)
            current_y += line_height + 2

    def _wrap_text(self, draw, text: str, font, max_width: int) -> list[str]:
        normalized = " ".join((text or "").split()).strip()
        if not normalized:
            return [""]

        words = normalized.split()
        lines: list[str] = []
        current = words[0]
        for word in words[1:]:
            candidate = f"{current} {word}"
            if draw.textbbox((0, 0), candidate, font=font)[2] <= max_width:
                current = candidate
            else:
                lines.append(current)
                current = word
        lines.append(current)
        return lines[:3]

    def _text_block_height(self, draw, lines: list[str], font) -> int:
        height = 0
        for line in lines:
            bbox = draw.textbbox((0, 0), line, font=font)
            height += (bbox[3] - bbox[1]) + 2
        return max(0, height - 2)

    def _load_font(self, size: int, *, bold: bool) -> object:
        font_name = "calibrib.ttf" if bold else "calibri.ttf"
        font_path = Path("C:/Windows/Fonts") / font_name
        try:
            return ImageFont.truetype(str(font_path), size)
        except Exception:
            return ImageFont.load_default()

    def _require_source_sheet(self):
        if self.source_sheet is None:
            raise RuntimeError("El Excel no está abierto.")
        return self.source_sheet

    @staticmethod
    def _normalize_excel_text(value) -> str:
        return normalize_excel_scalar(value)

    @staticmethod
    def _normalize_excel_identifier(value) -> str:
        return normalize_excel_scalar(value)

    @staticmethod
    def _pil_resize_filter():
        return getattr(getattr(Image, "Resampling", Image), "LANCZOS")
