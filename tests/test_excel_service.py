import unittest
from pathlib import Path
from tempfile import NamedTemporaryFile
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from io import BytesIO
from PIL import Image as PILImage
from PySide6.QtGui import QImage, QGuiApplication

from services.excel_service import ExcelService


class TestExcelService(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.app = QGuiApplication.instance() or QGuiApplication([])

    def test_extract_text_and_images(self):
        # Create a temporary excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Hoja1"

        # Add headers (not strictly necessary but good for context, let's assume standard extraction skips header or reads specific columns)
        ws.cell(row=1, column=1, value="ID")
        ws.cell(row=1, column=2, value="Name")
        ws.cell(row=1, column=3, value="Section")

        # Add data
        ws.cell(row=2, column=1, value="A-001")
        ws.cell(row=2, column=2, value="Laptop")
        ws.cell(row=2, column=3, value="IT")

        ws.cell(row=3, column=1, value="A-002")
        ws.cell(row=3, column=2, value="Monitor")
        ws.cell(row=3, column=3, value="HR")

        # Add an image to row 2
        # Create a small dummy image in memory
        img_io = BytesIO()
        pil_img = PILImage.new("RGB", (10, 10), color="red")
        pil_img.save(img_io, format="PNG")
        img_io.seek(0)

        img = OpenpyxlImage(img_io)
        # Anchor the image to row 2, column 4 (D2)
        # Note: openpyxl anchor coordinates are 0-indexed: col, row. So row 2 is index 1.
        img.anchor = openpyxl.drawing.spreadsheet_drawing.OneCellAnchor(
            _from=openpyxl.drawing.spreadsheet_drawing.AnchorMarker(col=3, row=1)
        )
        ws.add_image(img)

        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb.save(tmp.name)
            self.temp_file = tmp.name

        service = ExcelService()
        records = service.extract_data(self.temp_file)

        self.assertEqual(len(records), 2)

        # Check record 1
        self.assertEqual(records[0].asset_id, "A-001")
        self.assertEqual(records[0].asset_name, "Laptop")
        self.assertEqual(records[0].section, "IT")
        self.assertIsNotNone(records[0].image)
        self.assertIsInstance(records[0].image, QImage)

        # Check record 2
        self.assertEqual(records[1].asset_id, "A-002")
        self.assertIsNone(records[1].image)

        Path(self.temp_file).unlink()


if __name__ == "__main__":
    unittest.main()
